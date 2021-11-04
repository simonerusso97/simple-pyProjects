# This is a sample Python script.

# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.

# Press the green button in the gutter to run the script.
import pandas as pd
import xlrd
from openpyxl import load_workbook
from dataclasses import dataclass
from typing import List

giornataCorrente=14

def prendiPunti(squadra):
    return squadra.punti

@dataclass
class Squadra:
    id: str
    punti: int
    vittorie: int
    pareggi: int
    sconfitte: int

    vittorieInCasa: int
    pareggiInCasa: int
    sconfitteInCasa: int

    vittorieInTrasferta: int
    pareggiInTrasferta: int
    sconfitteInTrasferta: int

    retiSegnate: int
    retiSubite: int

    retiSegnateInCasa: int
    retiSubiteInCasa: int

    retiSegnateInTraferta: int
    retiSubiteInTraferta: int

    serie: []
    serieInCasa: []
    serieInTrasferta: []

filename = (r'/Users/simonerusso/Desktop/serieA.xlsx')
workbook = load_workbook(filename=filename)
workbook.sheetnames
sheet = workbook.active


max_row=(11*(giornataCorrente-1))
C = {}

squadre = [Squadra('Atalanta',0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,[], [], [] ), Squadra('Benevento',0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,[], [], []), Squadra('Bologna',0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,[], [], []), Squadra('Cagliari',0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,[], [], []), Squadra('Crotone',0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,[], [], []), Squadra('Fiorentina',0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,[], [], []), Squadra('Genoa',0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,[], [], []), Squadra('Inter',0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,[], [], []), Squadra('Juventus',0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,[], [], []), Squadra('Lazio',0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,[], [], []), Squadra('Milan',0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,[], [], []), Squadra('Parma',0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,[], [], []), Squadra('Roma',0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,[], [], []), Squadra('Sampdoria',0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,[], [], []), Squadra('Sassuolo',0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,[], [], []), Squadra('Spezia',0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,[], [], []), Squadra('Torino',0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,[], [], []), Squadra('Udinese',0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,[], [], []), Squadra('Verona',0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,[], [], []), Squadra('Napoli',0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,[], [], [])]
for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=2, values_only=True):

    if(type(row[0]) != int and row[1]!='/'):
        squadraCasa = row[0].split("-")[0]
        squadraOspite = row[0].split("-")[1]
        golSquadraCasa=int(row[1].split("-")[0])
        golSquadraOspite = int(row[1].split("-")[1])
        C[(squadraCasa, squadraOspite)] = row[1]
        if ( golSquadraCasa > golSquadraOspite):
            esito="1"
        elif (golSquadraCasa < golSquadraOspite):
            esito="2"
        else:
            esito="x"




        for squadra in squadre:
            if (squadra.id == squadraCasa):
                squadra.retiSegnate += golSquadraCasa
                squadra.retiSegnateInCasa += golSquadraCasa
                squadra.retiSubite+=golSquadraOspite
                squadra.retiSubiteInCasa+=golSquadraOspite

                if(esito=="1"):
                    squadra.vittorie+=1
                    squadra.vittorieInCasa+=1
                    squadra.serie.append("v")
                    squadra.serieInCasa.append("v")

                elif(esito=="2"):
                    squadra.sconfitte+=1
                    squadra.sconfitteInCasa+=1
                    squadra.serie.append("s")
                    squadra.serieInCasa.append("s")

                elif(esito=="x"):
                    squadra.pareggi+=1
                    squadra.pareggiInCasa+=1
                    squadra.serie.append("x")
                    squadra.serieInCasa.append("x")

            elif (squadra.id == squadraOspite):
                squadra.retiSegnate += golSquadraOspite
                squadra.retiSegnateInTraferta += golSquadraOspite
                squadra.retiSubite+=golSquadraCasa
                squadra.retiSubiteInTraferta+=golSquadraCasa
                if(esito=="1"):
                    squadra.sconfitte+=1
                    squadra.sconfitteInTrasferta+=1
                    squadra.serie.append("s")
                    squadra.serieInTrasferta.append("s")


                elif(esito=="2"):
                    squadra.vittorie+=1
                    squadra.vittorieInTrasferta+=1
                    squadra.serie.append("v")
                    squadra.serieInTrasferta.append("v")

                elif(esito=="x"):
                    squadra.pareggi+=1
                    squadra.pareggiInTrasferta+=1
                    squadra.serie.append("x")
                    squadra.serieInTrasferta.append("x")
for squadra in squadre:
    squadra.punti+=(3*squadra.vittorie)+squadra.pareggi
squadre.sort(key=prendiPunti)

min_row = max_row+1
max_row+=11

zonaDiClassificaSquadraCasa = []
zonaDiClassificaSquadraTrasferta = []
#CALCOLO DELLE PROBABILITA'
for row in sheet.iter_rows(min_row, max_row=max_row, min_col=1, max_col=2, values_only=True):
    if (type(row[0]) != int):
        squadraCasa = row[0].split("-")[0]
        squadraOspite = row[0].split("-")[1]
        pr1=0
        pr2=0
        prx=0

        for i in squadre:
            for j in squadre:
                if (j.id == squadraCasa):
                    break

            if(i != j and i.punti >= (j.punti - 5) and (j.punti + 5) >= i.punti):
                zonaDiClassificaSquadraCasa.append(i)
            for j in squadre:
                if (j.id == squadraOspite):
                    break
            if(i != j and i.punti >= (j.punti - 5) and (j.punti + 5) >= i.punti):
                zonaDiClassificaSquadraTrasferta.append(i)


        checkSquadraCasa=False
        checkSquadraOspite=False
        for squadra in squadre:
            if(squadra.id==squadraCasa):
                pr1+=squadra.vittorie+squadra.vittorieInCasa
                prx+=squadra.pareggi+squadra.pareggiInCasa
                pr2+=squadra.sconfitte+squadra.sconfitteInCasa
                for i in squadra.serieInCasa:
                    if(i=="v"):
                        pr1+=1
                    elif(i=="s"):
                        pr2+=1
                    else:
                        prx+=1
                for i in squadra.serie:
                    if(i=="v"):
                        pr1+=1
                    elif(i=="s"):
                        pr2+=1
                    else:
                        prx+=1
                for i in zonaDiClassificaSquadraTrasferta:
                    if(i!=squadraCasa):
                        try:
                            risultato = C[(i, squadraOspite)]
                            if(risultato.split("-")[0] > risultato.split("-")[1]):
                                pr1 += 1
                            elif(risultato.split("-")[0] < risultato.split("-")[1]):
                                pr2+=1
                            else:
                                prx+=1
                        except:
                            try:
                                risultato = C[(squadraOspite, i)]
                                if (risultato.split("-")[0] < risultato.split("-")[1]):
                                    pr1 += 1
                                elif (risultato.split("-")[0] > risultato.split("-")[1]):
                                    pr2 += 1
                                else:
                                    prx += 1
                            except:
                                continue



                checkSquadraCasa=True
            if (squadra.id == squadraOspite):
                pr1 += squadra.sconfitte + squadra.sconfitteInTrasferta
                prx += squadra.pareggi + squadra.pareggiInTrasferta
                pr2 += squadra.vittorie + squadra.vittorieInTrasferta
                for i in squadra.serieInTrasferta:
                    if (i == "v"):
                        pr2 += 1
                    elif (i == "s"):
                        pr1 += 1
                    else:
                        prx += 1
                for i in squadra.serie:
                    if (i == "v"):
                        pr2 += 1
                    elif (i == "s"):
                        pr1 += 1
                    else:
                        prx += 1
                for i in zonaDiClassificaSquadraTrasferta:
                    try:
                        if(i!=squadraCasa):
                            risultato = C[(squadraCasa, i)]
                            if(risultato.split("-")[0] > risultato.split("-")[1]):
                                pr1 += 1
                            elif(risultato.split("-")[0] < risultato.split("-")[1]):
                                pr2+=1
                            else:
                                prx+=1
                    except:
                        try:
                            if (i != squadraCasa):
                                risultato = C[(i, squadraCasa)]
                                if (risultato.split("-")[0] < risultato.split("-")[1]):
                                    pr1 += 1
                                elif (risultato.split("-")[0] > risultato.split("-")[1]):
                                    pr2 += 1
                                else:
                                    prx += 1
                        except:
                            continue
                checkSquadraOspite=True
            if(checkSquadraOspite==True and checkSquadraCasa==True):
                tot=pr1+prx+pr2
                pr1=int((pr1*10000)/tot)/100
                prx=int((prx*10000)/tot)/100
                pr2=int((pr2*10000)/tot)/100

                print(squadraCasa+"-"+squadraOspite)
                print("1: "+str(pr1)+"%")
                print("x: "+str(prx)+"%")
                print("2: "+str(pr2)+"%")
                print()

                break

        zonaDiClassificaSquadraCasa.clear()
        zonaDiClassificaSquadraTrasferta.clear()